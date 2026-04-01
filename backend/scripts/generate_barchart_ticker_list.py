"""
generate_barchart_ticker_list.py
Gera planilha com tickers organizados para uso com Barchart for Excel.

ABORDAGEM: A formula =@BETS(...) NAO e gravada no arquivo porque o openpyxl
corrompe o XML do Excel ao tentar serializar formulas com prefixo @.
Em vez disso, a formula e exibida como texto puro na celula A4.
O usuario clica em A4, adiciona = no inicio da barra de formulas e pressiona Enter.

Uso:
    cd backend
    .venv\Scripts\python.exe -m scripts.generate_barchart_ticker_list
    .venv\Scripts\python.exe -m scripts.generate_barchart_ticker_list --batch-size 200 --bars 2520
"""

from __future__ import annotations

import argparse
import asyncio
import os
from datetime import date

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DB = os.environ.get(
    "DATABASE_URL_SYNC",
    "postgresql://tsdbadmin:LFBuQ5TEio_dP47w5P201NydByiw2RO3@nvhhm6dwvh.keh9pcdgv1.tsdb.cloud.timescale.com:30124/tsdb",
).replace("postgresql+psycopg://", "postgresql://").replace("postgresql+asyncpg://", "postgresql://")

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
TICKER_FONT  = Font(bold=True, size=10, color="1F4E79")
FORMULA_FONT = Font(bold=True, size=9, color="375623")
INSTR_FONT   = Font(size=9, italic=True, color="7F6000")

SOURCE_FILL = {
    "ETF":  PatternFill("solid", start_color="DDEEFF"),
    "MF":   PatternFill("solid", start_color="EEFFDD"),
    "ESMA": PatternFill("solid", start_color="FFEECC"),
}


async def fetch_tickers(conn: asyncpg.Connection) -> list[dict]:
    rows: list[dict] = []
    seen: set[str] = set()

    etfs = await conn.fetch(
        "SELECT ticker, 'ETF' AS source, fund_name AS name FROM sec_etfs "
        "WHERE ticker IS NOT NULL ORDER BY monthly_avg_net_assets DESC NULLS LAST, ticker"
    )
    for r in etfs:
        if r["ticker"] not in seen:
            seen.add(r["ticker"])
            rows.append(dict(r))

    mf = await conn.fetch(
        "SELECT DISTINCT ON (fc.series_id) fc.ticker, 'MF' AS source, fc.series_name AS name "
        "FROM sec_fund_classes fc WHERE fc.ticker IS NOT NULL "
        "ORDER BY fc.series_id, fc.expense_ratio_pct ASC NULLS LAST"
    )
    for r in mf:
        if r["ticker"] not in seen:
            seen.add(r["ticker"])
            rows.append(dict(r))

    esma = await conn.fetch(
        "SELECT yahoo_ticker AS ticker, 'ESMA' AS source, fund_name AS name "
        "FROM esma_funds WHERE yahoo_ticker IS NOT NULL ORDER BY fund_name"
    )
    for r in esma:
        if r["ticker"] not in seen:
            seen.add(r["ticker"])
            rows.append(dict(r))

    return rows


def write_batch_sheet(wb: Workbook, batch_num: int, rows: list[dict], bars: int) -> None:
    sources = list(dict.fromkeys(r["source"] for r in rows))
    sheet_name = f"Batch_{batch_num:02d}_({'_'.join(sources)})"[:31]
    ws = wb.create_sheet(title=sheet_name)

    n = len(rows)
    last_col = get_column_letter(n + 1)
    symbol_range = f"B1:{last_col}1"

    # ── Linha 1: "Time Series" + tickers ─────────────────────────────────────
    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1, value="Time Series")
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 55

    for col_idx, r in enumerate(rows, start=2):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=1, column=col_idx, value=r["ticker"])
        c.font = TICKER_FONT
        c.fill = SOURCE_FILL.get(r["source"], PatternFill("solid", start_color="FFFFFF"))
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[col_letter].width = 10

    # ── Linha 2: "Date" + "Close" ─────────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    ws.cell(row=2, column=1, value="Date").font = Font(bold=True, size=9)
    for col_idx in range(2, n + 2):
        c = ws.cell(row=2, column=col_idx, value="Close")
        c.font = Font(size=9, italic=True, color="595959")
        c.alignment = Alignment(horizontal="center")

    # ── Linha 3: instrucoes ───────────────────────────────────────────────────
    ws.row_dimensions[3].height = 40
    instr = (
        "INSTRUCOES: "
        "1) Clique na celula A4 abaixo (texto verde). "
        "2) Na BARRA DE FORMULAS no topo, clique no inicio do texto e digite = (sinal de igual). "
        "3) Pressione Enter. O Barchart ira popular os dados automaticamente. "
        "4) Aguarde o carregamento. "
        f"5) Salve como Batch_{batch_num:02d}_nav.xlsx em D:/Projetos/nav_batches/"
    )
    ws.merge_cells(f"A3:{get_column_letter(min(n + 1, 20))}3")
    c = ws.cell(row=3, column=1, value=instr)
    c.font = INSTR_FONT
    c.fill = PatternFill("solid", start_color="FFF2CC")
    c.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

    # ── Linha 4: formula como TEXTO PURO (sem = para evitar corrupcao XML) ────
    # O usuario adiciona = manualmente na barra de formulas
    # DATA(2015;1;1) = StartDate como 6o parametro (posicional, PT-BR)
    # Confirmado pelo usuario: esta sintaxe funciona para historico de 10 anos
    # Sort=Descending porque o Barchart retorna mais rapido assim — seed script ordena ASC
    formula_text = (
        f'@BETS({symbol_range};B2:B2;'
        f'"Aggregation=Day";"AggSize=1";"AggSpec=None";DATA(2015;1;1);;;;;"Sort=Descending";'
        f'"Orientation=Vertical";"HideWeekends=False";"Bars={bars}";'
        f'"Volume=Contract";"Tooltip=True")'
    )
    ws.row_dimensions[4].height = 20
    c = ws.cell(row=4, column=1, value=formula_text)
    c.font = FORMULA_FONT
    c.fill = PatternFill("solid", start_color="E2EFDA")
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A5"


def write_summary_sheet(wb: Workbook, all_rows: list[dict], batch_size: int, bars: int) -> None:
    ws = wb.create_sheet(title="LEIA_PRIMEIRO", index=0)
    ws.column_dimensions["A"].width = 70

    ws.cell(row=1, column=1, value="Barchart NAV Seed").font = Font(bold=True, size=14, color="1F4E79")
    ws.cell(row=2, column=1, value=f"Gerado: {date.today()}  |  {len(all_rows)} tickers  |  Bars={bars} (~{bars//252} anos)").font = Font(size=10, color="595959")

    steps = [
        ("", ""),
        ("COMO USAR:", ""),
        ("1.", "Abra uma aba Batch_XX neste arquivo."),
        ("2.", "Clique na celula A4 (texto verde com a formula BETS)."),
        ("3.", "Na BARRA DE FORMULAS (topo do Excel), clique antes do @ e digite = (sinal de igual)."),
        ("4.", "Pressione Enter. O Barchart popula os dados automaticamente."),
        ("5.", "Aguarde (200 tickers x 10 anos pode levar 2-5 minutos por aba)."),
        ("6.", "Salve como Batch_XX_nav.xlsx em D:/Projetos/nav_batches/"),
        ("7.", "Repita para cada aba Batch."),
        ("8.", "Quando todos os batches salvos, execute o seed script:"),
        ("",  "   cd D:\\Projetos\\netz-analysis-engine\\backend"),
        ("",  "   .venv\\Scripts\\python.exe -m scripts.seed_nav_barchart D:/Projetos/nav_batches/"),
        ("", ""),
        ("POR QUE TEXTO E NAO FORMULA?", "openpyxl corromperia o XML ao gravar =@BETS(). Adicionar = manualmente evita o problema."),
        ("LIMITE BARCHART EOD:", f"50.000 queries/dia — {len(all_rows)} tickers esta dentro do limite."),
    ]

    for i, (num, text) in enumerate(steps, start=4):
        c1 = ws.cell(row=i, column=1, value=num)
        c1.font = Font(bold=bool(num), size=10, color="1F4E79" if num else "333333")
        c2 = ws.cell(row=i, column=2, value=text)
        c2.font = Font(size=10, color="333333")
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)

    by_source: dict[str, int] = {}
    for r in all_rows:
        by_source[r["source"]] = by_source.get(r["source"], 0) + 1

    row = 4 + len(steps) + 1
    ws.cell(row=row, column=1, value="RESUMO:").font = Font(bold=True, size=10)
    row += 1
    for src, cnt in sorted(by_source.items()):
        ws.cell(row=row, column=1, value=f"  {src}").font = Font(size=10)
        ws.cell(row=row, column=2, value=cnt).font = Font(size=10)
        row += 1
    ws.cell(row=row, column=1, value="  TOTAL").font = Font(bold=True, size=10)
    ws.cell(row=row, column=2, value=len(all_rows)).font = Font(bold=True, size=10)


async def main(batch_size: int, output_path: str, bars: int) -> None:
    conn = await asyncpg.connect(DB, ssl="require")
    print("Buscando tickers do DB...")
    rows = await fetch_tickers(conn)
    await conn.close()

    print(f"Total tickers unicos: {len(rows)}")
    print(f"Bars: {bars} (~{bars//252} anos)")
    print(f"Batch size: {batch_size}")

    wb = Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, rows, batch_size, bars)

    batches = [rows[i: i + batch_size] for i in range(0, len(rows), batch_size)]
    for batch_num, batch in enumerate(batches, start=1):
        write_batch_sheet(wb, batch_num, batch, bars)
        sources = set(r["source"] for r in batch)
        print(f"  Batch {batch_num:02d}: {len(batch)} tickers ({', '.join(sorted(sources))})")

    wb.save(output_path)
    print(f"\nSalvo em: {output_path}")
    print(f"Total batches: {len(batches)}")
    print()
    print("PROXIMO PASSO:")
    print("  Abra no Excel com Barchart instalado.")
    print("  Cada aba tem a formula BETS como TEXTO em A4.")
    print("  Clique A4 > barra de formulas > adicione = no inicio > Enter.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch-size", type=int, default=200)
    parser.add_argument("--output", default="D:/Projetos/barchart_ticker_list.xlsx")
    parser.add_argument("--bars", type=int, default=2520,
                        help="Pregoes historicos (default 2520 = ~10 anos)")
    args = parser.parse_args()
    asyncio.run(main(args.batch_size, args.output, args.bars))
